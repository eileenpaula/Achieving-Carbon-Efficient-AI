import torch
import torchvision
import torchvision.transforms as transforms
import torch.nn as nn
import torchvision.models as models
import torch.optim as optim
from carbontracker.tracker import CarbonTracker
from carbontracker import parser
import time
import os
from openpyxl import load_workbook, Workbook
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score

if not os.path.exists("results.xlsx"):
    wb = Workbook()
    wb.save("results.xlsx")

transform = transforms.Compose([transforms.ToTensor(), transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5))])

trainset = torchvision.datasets.CIFAR10(root='./data', train=True, download=True, transform=transform)
trainloader = torch.utils.data.DataLoader(trainset, batch_size=128, shuffle=True, num_workers=2)
testset = torchvision.datasets.CIFAR10(root='./data', train=False, download=True, transform=transform)
testloader = torch.utils.data.DataLoader(testset, batch_size=128, shuffle=False, num_workers=2)

class ResNetCIFAR10(nn.Module):
    def __init__(self):
        super(ResNetCIFAR10, self).__init__()
        self.resnet = models.resnet18(pretrained=True)
        self.resnet.fc = nn.Linear(512, 10) 

    def forward(self, x):
        return self.resnet(x)

model = ResNetCIFAR10().cuda()
criterion = nn.CrossEntropyLoss()
optimizer = optim.SGD(model.parameters(), lr=0.01, momentum=0.9)
tracker = CarbonTracker(
    epochs=10,
    epochs_before_pred=-1,
    monitor_epochs=-1,
    interpretable=True,
    log_dir="./resnet_cifar10",
    verbose=2
)
start_time = time.time()

for epoch in range(10):
    tracker.epoch_start()
    running_loss = 0.0
    for i, data in enumerate(trainloader, 0):
        inputs, labels = data
        inputs, labels = inputs.cuda(), labels.cuda()
        optimizer.zero_grad()
        outputs = model(inputs)
        loss = criterion(outputs, labels)
        loss.backward()
        optimizer.step()
        running_loss += loss.item()
        if i % 100 == 99:   
            print(f'[Epoch: {epoch + 1}, Batch: {i + 1}] loss: {running_loss / 100:.3f}')
            running_loss = 0.0
    tracker.epoch_end()

tracker.stop()
training_time = time.time() - start_time

print('Finished Training')
y_true = []
y_pred = []

with torch.no_grad():
    for data in testloader:
        images, labels = data
        images, labels = images.cuda(), labels.cuda()
        outputs = model(images)
        _, predicted = torch.max(outputs.data, 1)
        y_true.extend(labels.cpu().numpy())
        y_pred.extend(predicted.cpu().numpy())

accuracy = accuracy_score(y_true, y_pred)
precision = precision_score(y_true, y_pred, average='weighted')
recall = recall_score(y_true, y_pred, average='weighted')
f1 = f1_score(y_true, y_pred, average='weighted')

# Parse logs
log_dir = "./resnet_cifar10"
logs = parser.parse_all_logs(log_dir=log_dir)

# Initialize variables to handle missing data
energy_kwh = 0.0
co2_eq_g = 0.0

first_log = logs[0]

resnet_data = {
    "Model": "ResNet",
    "Dataset": "CIFAR10",
    "Evaluation Framework": "carbontracker",
    "Total Energy (kWh)": first_log.get("actual", {}).get("energy (kWh)", 0.0),
    "Total CO2 Emissions (kgCO2e)": (first_log.get("actual", {}).get("co2eq (g)", 0.0)) / 1000,
    "Training Time (minutes)": training_time / 60,
    "Accuracy": accuracy,
    "Precision": precision,
    "Recall": recall,
    "F1": f1,
    "Number of Epochs": 10
}

wb = load_workbook("results.xlsx")
ws = wb.active
ws.append(["", ""])
for key, value in resnet_data.items():
    ws.append([key, value])

wb.save("results.xlsx")
print("ResNet data added to results.xlsx")