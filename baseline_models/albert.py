import torch
from transformers import AlbertTokenizer, AlbertForSequenceClassification, AdamW, get_scheduler
from datasets import load_dataset
from torch.utils.data import DataLoader
from tqdm.auto import tqdm
from codecarbon import EmissionsTracker
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_auc_score, confusion_matrix
from openpyxl import load_workbook, Workbook
import time
import os

# Check GPU availability
if not torch.cuda.is_available():
    raise RuntimeError("CUDA is not available. Please check your GPU configuration.")

# Initialize spreadsheet if it doesn't exist
if not os.path.exists("albertresults.xlsx"):
    wb = Workbook()
    wb.save("albertresults.xlsx")

# Load the Amazon Polarity dataset
dataset = load_dataset("amazon_polarity")
train_subset = dataset["train"].train_test_split(test_size=0.25)["train"]
test_dataset = dataset["test"]

# Define tokenizer and model
tokenizer = AlbertTokenizer.from_pretrained("albert-base-v2")
model = AlbertForSequenceClassification.from_pretrained("albert-base-v2", num_labels=2)

# Tokenization function
def tokenize_function(examples):
    return tokenizer(examples["content"], padding="max_length", truncation=True, max_length=128)

# Tokenize datasets
tokenized_train = train_subset.map(tokenize_function, batched=True, remove_columns=["content"])
tokenized_test = test_dataset.map(tokenize_function, batched=True, remove_columns=["content"])

# Rename label column and set format
tokenized_train = tokenized_train.rename_column("label", "labels")
tokenized_train.set_format(type="torch", columns=["input_ids", "attention_mask", "labels"])
tokenized_test = tokenized_test.rename_column("label", "labels")
tokenized_test.set_format(type="torch", columns=["input_ids", "attention_mask", "labels"])

# Create DataLoaders
batch_size = 16
train_loader = DataLoader(tokenized_train, shuffle=True, batch_size=batch_size)
test_loader = DataLoader(tokenized_test, batch_size=batch_size)

# Initialize optimizer and scheduler
device = torch.device("cuda")
model.to(device)
optimizer = AdamW(model.parameters(), lr=2e-5)
num_epochs = 3
num_training_steps = num_epochs * len(train_loader)
lr_scheduler = get_scheduler("linear", optimizer=optimizer, num_warmup_steps=0, num_training_steps=num_training_steps)

# Initialize emissions tracker
tracker = EmissionsTracker()
tracker.start()

# Training loop with evaluation at each epoch
progress_bar = tqdm(range(num_training_steps))
start_time = time.time()
for epoch in range(num_epochs):
    model.train()
    running_loss = 0.0
    for batch in train_loader:
        batch = {k: v.to(device) for k, v in batch.items()}
        optimizer.zero_grad()
        outputs = model(**batch)
        loss = outputs.loss
        loss.backward()
        optimizer.step()
        lr_scheduler.step()
        running_loss += loss.item()
        progress_bar.update(1)
    print(f"Epoch {epoch + 1}/{num_epochs}, Loss: {running_loss / len(train_loader):.4f}")

    # Evaluation at the end of the epoch
    model.eval()
    all_labels, all_predictions, all_probabilities = [], [], []
    with torch.no_grad():
        for batch in test_loader:
            batch = {k: v.to(device) for k, v in batch.items()}
            outputs = model(**batch)
            logits = outputs.logits
            predictions = torch.argmax(logits, dim=-1)
            probabilities = torch.nn.functional.softmax(logits, dim=-1)[:, 1]

            all_predictions.extend(predictions.cpu().numpy())
            all_labels.extend(batch["labels"].cpu().numpy())
            all_probabilities.extend(probabilities.cpu().numpy())

    # Compute metrics
    accuracy = accuracy_score(all_labels, all_predictions)
    precision = precision_score(all_labels, all_predictions, average="weighted")
    recall = recall_score(all_labels, all_predictions, average="weighted")
    f1 = f1_score(all_labels, all_predictions, average="weighted")
    roc_auc = roc_auc_score(all_labels, all_probabilities)
    cm = confusion_matrix(all_labels, all_predictions)

    print(f"Epoch {epoch + 1} Metrics: Accuracy {accuracy:.4f}, Precision {precision:.4f}, "
          f"Recall {recall:.4f}, F1 {f1:.4f}, ROC AUC {roc_auc:.4f}")
    print(f"Confusion Matrix:\n{cm}")

model.eval()
all_labels, all_predictions, all_probabilities = [], [], []
with torch.no_grad():
    for batch in test_loader:
        batch = {k: v.to(device) for k, v in batch.items()}
        outputs = model(**batch)
        logits = outputs.logits
        predictions = torch.argmax(logits, dim=-1)
        probabilities = torch.nn.functional.softmax(logits, dim=-1)[:, 1]

        all_predictions.extend(predictions.cpu().numpy())
        all_labels.extend(batch["labels"].cpu().numpy())
        all_probabilities.extend(probabilities.cpu().numpy())

# Compute metrics
accuracy = accuracy_score(all_labels, all_predictions)
precision = precision_score(all_labels, all_predictions, average="weighted")
recall = recall_score(all_labels, all_predictions, average="weighted")
f1 = f1_score(all_labels, all_predictions, average="weighted")
roc_auc = roc_auc_score(all_labels, all_probabilities)
cm = confusion_matrix(all_labels, all_predictions)

# Stop the tracker
tracker.stop()
training_time = (time.time() - start_time) / 60

print(f"Epoch {epoch + 1} Metrics: Accuracy {accuracy:.4f}, Precision {precision:.4f}, "
        f"Recall {recall:.4f}, F1 {f1:.4f}, ROC AUC {roc_auc:.4f}")
print(f"Confusion Matrix:\n{cm}")

# Prepare data for saving
albert_data = {
    "Model": "ALBERT Baseline",
    "Dataset": "Amazon Polarity",
    "Evaluation Framework": "codecarbon",
    "Total Energy (kWh)": tracker.final_emissions_data.energy_consumed,
    "Total CO2 Emissions (kgCO2e)": tracker.final_emissions_data.emissions,
    "CPU Energy": tracker.final_emissions_data.cpu_energy,
    "GPU Energy": tracker.final_emissions_data.gpu_energy,
    "RAM Energy": tracker.final_emissions_data.ram_energy,
    "Emissions Rate": tracker.final_emissions_data.emissions_rate,
    "Training Time (minutes)": training_time,
    "Accuracy": accuracy,
    "Precision": precision,
    "Recall": recall,
    "F1": f1,
    "ROC AUC": roc_auc,
    "Confusion Matrix": "refer to terminal",
    "Number of Epochs": num_epochs
}

# Write data to Excel
wb = load_workbook("albertresults.xlsx")
ws = wb.active
ws.append(["", ""])
for key, value in albert_data.items():
    ws.append([key, value])

wb.save("albertresults.xlsx")
print("Albert Baseline data added to albertresults.xlsx")