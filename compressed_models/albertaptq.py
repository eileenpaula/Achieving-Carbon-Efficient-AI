import torch
from transformers import AlbertTokenizer, AlbertForSequenceClassification, AdamW, get_scheduler
from datasets import load_dataset
from torch.utils.data import DataLoader
from tqdm.auto import tqdm
from codecarbon import EmissionsTracker
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_auc_score, confusion_matrix
from openpyxl import load_workbook, Workbook
import time
import os
import openvino as ov
import nncf
from nncf.parameters import ModelType
import numpy as np

# Ensure CUDA is available for training
if not torch.cuda.is_available():
    raise RuntimeError("CUDA is not available. Please check your GPU configuration.")

# Initialize spreadsheet if it doesn't exist
if not os.path.exists("results.xlsx"):
    wb = Workbook()
    wb.save("results.xlsx")

# Load Amazon Polarity dataset
dataset = load_dataset("amazon_polarity")

# Split training set (keep 75% for training)
train_subset = dataset["train"].train_test_split(test_size=0.25)["train"]
test_dataset = dataset["test"]

# Define tokenizer and ALBERT model
tokenizer = AlbertTokenizer.from_pretrained("albert-base-v2")
model_fp32 = AlbertForSequenceClassification.from_pretrained("albert-base-v2", num_labels=2)

# Tokenization function
def tokenize_function(examples):
    return tokenizer(examples["content"], padding="max_length", truncation=True, max_length=128)

# Tokenize train and test datasets
tokenized_train = train_subset.map(tokenize_function, batched=True, remove_columns=["content"])
tokenized_test = test_dataset.map(tokenize_function, batched=True, remove_columns=["content"])

# Rename "label" to "labels" and format to PyTorch tensors
tokenized_train = tokenized_train.rename_column("label", "labels")
tokenized_train.set_format(type="torch", columns=["input_ids", "attention_mask", "labels"])
tokenized_test = tokenized_test.rename_column("label", "labels")
tokenized_test.set_format(type="torch", columns=["input_ids", "attention_mask", "labels"])

# Create DataLoaders
batch_size = 16
train_loader = DataLoader(tokenized_train, shuffle=True, batch_size=batch_size)
test_loader = DataLoader(tokenized_test, batch_size=batch_size)

# Check if GPU is available for training
device = torch.device("cuda")
model_fp32.to(device)

# Initialize optimizer and scheduler
optimizer = AdamW(model_fp32.parameters(), lr=2e-5)
num_epochs = 3
num_training_steps = num_epochs * len(train_loader)
lr_scheduler = get_scheduler(
    name="linear", optimizer=optimizer, num_warmup_steps=0, num_training_steps=num_training_steps
)

# Initialize CodeCarbon tracker for training
tracker = EmissionsTracker()
tracker.start()

# Training loop
progress_bar = tqdm(range(num_training_steps))
start_time = time.time()

model_fp32.train()
for epoch in range(num_epochs):
    running_loss = 0.0
    for batch in train_loader:
        batch = {k: v.to(device) for k, v in batch.items()}
        
        optimizer.zero_grad()
        outputs = model_fp32(**batch)
        loss = outputs.loss
        loss.backward()
        
        optimizer.step()
        lr_scheduler.step()
        running_loss += loss.item()
        progress_bar.update(1)

    print(f"Epoch {epoch + 1}/{num_epochs}, Loss: {running_loss / len(train_loader):.4f}")

# Move model to CPU for OpenVINO Quantization
model_fp32.to("cpu")

# Convert PyTorch model to OpenVINO IR (FP32)
ir_model_xml = "albert_fp32.xml"
core = ov.Core()

if not os.path.exists(ir_model_xml):
    input_shape = ov.PartialShape([1, 128])
    inputs = {
        "input_ids": torch.ones(1, 128, dtype=torch.int64),
        "attention_mask": torch.ones(1, 128, dtype=torch.int64)
    }
    ov_model = ov.convert_model(model_fp32, example_input=inputs, input=[
        ("input_ids", input_shape, torch.int64),
        ("attention_mask", input_shape, torch.int64)
    ])
    ov.save_model(ov_model, ir_model_xml)
else:
    ov_model = core.read_model(ir_model_xml)

# Prepare dataset for quantization
def transform_fn(data_item):
    return {k: torch.unsqueeze(torch.tensor(data_item[k]), 0) for k in ["input_ids", "attention_mask"]}

calibration_dataset = nncf.Dataset(tokenized_train, transform_fn)

# Apply NNCF Post-Training Quantization (PTQ)
quantized_model = nncf.quantize(ov_model, calibration_dataset, model_type=ModelType.TRANSFORMER)

# Save Quantized Model
compressed_model_xml = "albert_int8.xml"
ov.save_model(quantized_model, compressed_model_xml)

# Load OpenVINO quantized model for inference
compiled_quantized_model = core.compile_model(model=quantized_model, device_name="CPU")

# Retrieve correct input names dynamically
input_names = {input.any_name: input.any_name for input in compiled_quantized_model.inputs}

# Run inference
all_labels, all_predictions, all_probabilities = [], [], []
output_layer = compiled_quantized_model.outputs[0]  # Get model output layer

for batch in test_loader:
    for i in range(batch["input_ids"].shape[0]):  # Iterate through batch samples
        sample = {
            input_names["input_ids"]: batch["input_ids"][i].unsqueeze(0).numpy(),  # Ensure shape [1, 128]
            input_names["73"]: batch["attention_mask"][i].unsqueeze(0).numpy()  # Fix renaming issue
        }
        
        result = compiled_quantized_model(sample)[output_layer]

        prediction = np.argmax(result, axis=-1)[0]  # Get prediction
        probability = torch.nn.functional.softmax(torch.tensor(result), dim=-1)[:, 1].numpy()[0]

        all_predictions.append(prediction)
        all_labels.append(batch["labels"][i].numpy())
        all_probabilities.append(probability)

# Compute final metrics
accuracy = accuracy_score(all_labels, all_predictions)
precision = precision_score(all_labels, all_predictions, average="weighted")
recall = recall_score(all_labels, all_predictions, average="weighted")
f1 = f1_score(all_labels, all_predictions, average="weighted")
roc_auc = roc_auc_score(all_labels, all_probabilities)
cm = confusion_matrix(all_labels, all_predictions)

# Stop CodeCarbon tracker after inference
tracker.stop()
training_time = (time.time() - start_time) / 60  # in minutes

# Print results
print(f"Final Metrics - Accuracy: {accuracy:.4f}, Precision: {precision:.4f}, "
      f"Recall: {recall:.4f}, F1: {f1:.4f}, ROC AUC: {roc_auc:.4f}")
print(f"Confusion Matrix:\n{cm}")

# Save Results to Excel
albert_data = {
    "Model": "ALBERT with OpenVINO PTQ",
    "Dataset": "Amazon Polarity",
    "Evaluation Framework": "codecarbon",
    "Total Energy (kWh)": tracker.final_emissions_data.energy_consumed,
    "Total CO2 Emissions (kgCO2e)": tracker.final_emissions_data.emissions,
    "CPU Energy": tracker.final_emissions_data.cpu_energy,
    "GPU Energy": tracker.final_emissions_data.gpu_energy,
    "RAM Energy": tracker.final_emissions_data.ram_energy,
    "Emissions Rate": tracker.final_emissions_data.emissions_rate,
    "Training Time (minutes)": training_time,
    "Accuracy": accuracy,
    "Precision": precision,
    "Recall": recall,
    "F1": f1,
    "ROC AUC": roc_auc,
    "Confusion Matrix": "refer to terminal",
    "Number of Epochs": num_epochs
}

wb = load_workbook("results.xlsx")
ws = wb.active
ws.append(["", ""])
for key, value in albert_data.items():
    ws.append([key, value])

wb.save("results.xlsx")
print("ALBERT with OpenVINO PTQ data added to results.xlsx")
