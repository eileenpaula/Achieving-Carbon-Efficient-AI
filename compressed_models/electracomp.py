import torch
from transformers import ElectraTokenizer, ElectraForSequenceClassification, DistilBertForSequenceClassification, AdamW, get_scheduler
from datasets import load_dataset
from torch.utils.data import DataLoader
from tqdm.auto import tqdm
from codecarbon import EmissionsTracker
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_auc_score, confusion_matrix
from openpyxl import load_workbook, Workbook
import torch.nn.utils.prune as prune
import time
import os

if not torch.cuda.is_available():
    raise RuntimeError("CUDA is not available. Please check your GPU configuration.")

# Initialize spreadsheet if it doesn't exist
if not os.path.exists("electracompresults.xlsx"):
    wb = Workbook()
    wb.save("electracompresults.xlsx")

# Load the Amazon Polarity dataset
dataset = load_dataset("amazon_polarity")

# Split off a 25% sample of the training set
train_subset = dataset["train"].train_test_split(test_size=0.25)["train"]
test_dataset = dataset["test"]

# Define tokenizer and models
tokenizer = ElectraTokenizer.from_pretrained("google/electra-base-discriminator")
teacher_model = ElectraForSequenceClassification.from_pretrained("google/electra-base-discriminator", num_labels=2)  # Teacher (pruned ELECTRA)
student_model = DistilBertForSequenceClassification.from_pretrained("distilbert-base-uncased", num_labels=2)  # Student

# Tokenization function
def tokenize_function(examples):
    return tokenizer(examples["content"], padding="max_length", truncation=True, max_length=128)

# Tokenize train and test datasets
tokenized_train = train_subset.map(tokenize_function, batched=True)
tokenized_test = test_dataset.map(tokenize_function, batched=True)

# Rename "label" to "labels" and format to PyTorch tensors
tokenized_train = tokenized_train.rename_column("label", "labels")
tokenized_train.set_format(type="torch", columns=["input_ids", "attention_mask", "labels"])
tokenized_test = tokenized_test.rename_column("label", "labels")
tokenized_test.set_format(type="torch", columns=["input_ids", "attention_mask", "labels"])

# Create DataLoaders
batch_size = 16
train_loader = DataLoader(tokenized_train, shuffle=True, batch_size=batch_size)
test_loader = DataLoader(tokenized_test, batch_size=batch_size)

# Initialize CodeCarbon tracker to measure energy consumption
tracker = EmissionsTracker(allow_multiple_runs=True)
tracker.start()

# Move models to GPU
device = torch.device("cuda")
teacher_model.to(device)
student_model.to(device)

# Apply pruning to the teacher model
def apply_pruning(model, amount=0.1):
    for name, module in model.named_modules():
        if isinstance(module, torch.nn.Linear):
            # Make weights contiguous before pruning
            module.weight = torch.nn.Parameter(module.weight.contiguous())
            prune.l1_unstructured(module, name='weight', amount=amount)
            prune.remove(module, 'weight')  # Permanently apply pruning

apply_pruning(teacher_model, amount=0.1)

# Initialize optimizer and scheduler for the student model
optimizer = AdamW(student_model.parameters(), lr=1e-5)
num_epochs = 2
alpha = 0.5  # Weight for the loss combination
temperature = 6
num_training_steps = num_epochs * len(train_loader)
lr_scheduler = get_scheduler(
    name="linear", optimizer=optimizer, num_warmup_steps=0, num_training_steps=num_training_steps
)

start_time = time.time()

# Warm-up phase for the student model
warm_up_epochs = 1
for epoch in range(warm_up_epochs):
    student_model.train()
    running_loss = 0.0
    for batch in train_loader:
        batch = {k: v.to(device) for k, v in batch.items()}
        outputs = student_model(**batch)
        loss = torch.nn.functional.cross_entropy(outputs.logits, batch['labels'])
        
        loss.backward()
        optimizer.step()
        optimizer.zero_grad()
        running_loss += loss.item()
    print(f"Warm-up Epoch {epoch + 1}/{warm_up_epochs}, Loss: {running_loss / len(train_loader):.4f}")

# Distillation training with pruned teacher and real-time metrics logging
for epoch in range(num_epochs):
    teacher_model.eval()
    student_model.train()
    running_loss = 0.0
    all_labels = []
    all_predictions = []
    all_probabilities = []

    for batch in train_loader:
        batch = {k: v.to(device) for k, v in batch.items()}

        # Teacher model predictions (no grad)
        with torch.no_grad():
            teacher_outputs = teacher_model(**batch)
        teacher_logits = teacher_outputs.logits / temperature

        # Student model predictions
        student_outputs = student_model(**batch)
        student_logits = student_outputs.logits / temperature

        # Distillation loss (KL Divergence)
        kl_loss = torch.nn.functional.kl_div(
            torch.nn.functional.log_softmax(student_logits, dim=-1),
            torch.nn.functional.softmax(teacher_logits, dim=-1),
            reduction="batchmean"
        ) * (temperature ** 2)

        # Cross-Entropy loss with ground truth
        ce_loss = torch.nn.functional.cross_entropy(student_outputs.logits, batch["labels"])

        # Combined loss
        loss = alpha * kl_loss + (1 - alpha) * ce_loss

        # Backward and optimization step
        loss.backward()
        optimizer.step()
        optimizer.zero_grad()
        lr_scheduler.step()

        running_loss += loss.item()

        # Gather predictions and labels for real-time metrics logging
        predictions = torch.argmax(student_outputs.logits, dim=-1)
        # Apply softmax and select probabilities for the positive class
        probabilities = torch.nn.functional.softmax(outputs.logits, dim=-1)
        # Ensure probabilities has the right dimensions
        if probabilities.dim() == 2:
            all_probabilities.extend(probabilities[:, 1].detach().cpu().numpy())  # For 2D tensors
        else:
            all_probabilities.extend(probabilities.detach().cpu().numpy())  # For 1D tensors
        all_predictions.extend(predictions.cpu().numpy())
        all_labels.extend(batch["labels"].cpu().numpy())

    # Calculate and print metrics for this epoch
    accuracy = accuracy_score(all_labels, all_predictions)
    precision = precision_score(all_labels, all_predictions, average="weighted")
    recall = recall_score(all_labels, all_predictions, average="weighted")
    f1 = f1_score(all_labels, all_predictions, average="weighted")
    roc_auc = roc_auc_score(all_labels, all_probabilities)
    cm = confusion_matrix(all_labels, all_predictions)
    print(f"Distillation Epoch {epoch + 1}/{num_epochs}, Loss: {running_loss / len(train_loader):.4f}")
    print(f"Accuracy: {accuracy:.4f}, Precision: {precision:.4f}, Recall: {recall:.4f}, F1: {f1:.4f}, ROC AUC: {roc_auc:.4f}, Confusion Matrix: \n{cm}\n\n")

# Final evaluation on the test set
student_model.eval()
all_labels = []
all_predictions = []
all_probabilities = []
for batch in test_loader:
    batch = {k: v.to(device) for k, v in batch.items()}
    with torch.no_grad():
        outputs = student_model(**batch)
    predictions = torch.argmax(outputs.logits, dim=-1)
    # Apply softmax and select probabilities for the positive class
    probabilities = torch.nn.functional.softmax(outputs.logits, dim=-1)
    # Ensure probabilities has the right dimensions
    if probabilities.dim() == 2:
        all_probabilities.extend(probabilities[:, 1].detach().cpu().numpy())  # For 2D tensors
    else:
        all_probabilities.extend(probabilities.detach().cpu().numpy())  # For 1D tensors
    all_predictions.extend(predictions.cpu().numpy())
    all_labels.extend(batch["labels"].cpu().numpy())

# Compute final metrics
accuracy = accuracy_score(all_labels, all_predictions)
precision = precision_score(all_labels, all_predictions, average="weighted")
recall = recall_score(all_labels, all_predictions, average="weighted")
f1 = f1_score(all_labels, all_predictions, average="weighted")
roc_auc = roc_auc_score(all_labels, all_probabilities)
cm = confusion_matrix(all_labels, all_predictions)

print(f"Accuracy: {accuracy:.4f}, Precision: {precision:.4f}, Recall: {recall:.4f}, F1: {f1:.4f}, ROC AUC: {roc_auc:.4f}, Confusion Matrix: \n{cm}\n\n")

# Stop the CodeCarbon tracker and calculate training time
tracker.stop()
training_time = (time.time() - start_time) / 60  # in minutes

# Prepare data for saving
electra_distillation_data = {
    "Model": "ELECTRA with Pruning & Distillation",
    "Dataset": "Amazon Polarity",
    "Evaluation Framework": "codecarbon",
    "Total Energy (kWh)": tracker.final_emissions_data.energy_consumed,
    "Total CO2 Emissions (kgCO2e)": tracker.final_emissions_data.emissions,
    "CPU Energy": tracker.final_emissions_data.cpu_energy,
    "GPU Energy": tracker.final_emissions_data.gpu_energy,
    "RAM Energy": tracker.final_emissions_data.ram_energy,
    "Emissions Rate": tracker.final_emissions_data.emissions_rate,
    "Training Time (minutes)": training_time,
    "Accuracy": accuracy,
    "Precision": precision,
    "Recall": recall,
    "F1": f1,
    "ROC AUC": roc_auc,
    "Confusion Matrix": "refer to terminal",
    "Number of Epochs": num_epochs
}

# Write data to Excel
wb = load_workbook("electracompresults.xlsx")
ws = wb.active
ws.append(["", ""])
for key, value in electra_distillation_data.items():
    ws.append([key, value])

wb.save("electracompresults.xlsx")
print("ELECTRA with Pruning & Distillation data added to electracompresults.xlsx")
