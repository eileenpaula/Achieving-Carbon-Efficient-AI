import torch
from transformers import DistilBertTokenizer, DistilBertForSequenceClassification, AdamW, get_scheduler
from datasets import load_dataset
from torch.utils.data import DataLoader
from tqdm.auto import tqdm
from codecarbon import EmissionsTracker
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_auc_score, confusion_matrix
from openpyxl import load_workbook, Workbook
import torch.nn.utils.prune as prune
import time
import os

if not torch.cuda.is_available():
    raise RuntimeError("CUDA is not available. Please check your GPU configuration.")

# Initialize spreadsheet if it doesn't exist
if not os.path.exists("distilbertcompresults.xlsx"):
    wb = Workbook()
    wb.save("distilbertcompresults.xlsx")

# Load the full Amazon Polarity dataset
dataset = load_dataset("amazon_polarity")

# Split off a 25% sample of the training set
train_subset = dataset["train"].train_test_split(test_size=0.25)["train"]
test_dataset = dataset["test"]

# Define tokenizer and DistilBERT model
tokenizer = DistilBertTokenizer.from_pretrained("distilbert-base-uncased")
model = DistilBertForSequenceClassification.from_pretrained("distilbert-base-uncased", num_labels=2)

# Tokenization function
def tokenize_function(examples):
    return tokenizer(examples["content"], padding="max_length", truncation=True, max_length=128)

# Tokenize train and test datasets
tokenized_train = train_subset.map(tokenize_function, batched=True)
tokenized_test = test_dataset.map(tokenize_function, batched=True)

# Rename "label" to "labels" and format to PyTorch tensors
tokenized_train = tokenized_train.rename_column("label", "labels")
tokenized_train.set_format(type="torch", columns=["input_ids", "attention_mask", "labels"])
tokenized_test = tokenized_test.rename_column("label", "labels")
tokenized_test.set_format(type="torch", columns=["input_ids", "attention_mask", "labels"])

# Create DataLoaders
batch_size = 16
train_loader = DataLoader(tokenized_train, shuffle=True, batch_size=batch_size)
test_loader = DataLoader(tokenized_test, batch_size=batch_size)

# Initialize CodeCarbon tracker to measure energy consumption
tracker = EmissionsTracker(allow_multiple_runs=True)
tracker.start()

# Check if GPU is available
device = torch.device("cuda")
model.to(device)

# Apply pruning to the DistilBERT model
def apply_pruning(model, amount=0.1):
    # Apply pruning to each linear layer
    for name, module in model.named_modules():
        if isinstance(module, torch.nn.Linear):
            prune.l1_unstructured(module, name='weight', amount=amount)
            prune.remove(module, 'weight')  # Permanently apply pruning

apply_pruning(model, amount=0.1)

# Initialize optimizer and scheduler
optimizer = AdamW(model.parameters(), lr=2e-5)
num_epochs = 3
num_training_steps = num_epochs * len(train_loader)
lr_scheduler = get_scheduler(
    name="linear", optimizer=optimizer, num_warmup_steps=0, num_training_steps=num_training_steps
)

# Training loop with progress and loss tracking
progress_bar = tqdm(range(num_training_steps))
start_time = time.time()

model.train()

for epoch in range(num_epochs):
    running_loss = 0.0
    for batch in train_loader:
        batch = {k: v.to(device) for k, v in batch.items()}
        
        outputs = model(**batch)
        loss = outputs.loss
        loss.backward()
        
        optimizer.step()
        lr_scheduler.step()
        optimizer.zero_grad()
        progress_bar.update(1)
        
        running_loss += loss.item()

    print(f"Epoch {epoch + 1}/{num_epochs}, Loss: {running_loss / len(train_loader):.4f}")

    # Evaluation on test set at the end of each epoch
    model.eval()
    all_labels = []
    all_predictions = []
    all_probabilities = []
    for batch in test_loader:
        batch = {k: v.to(device) for k, v in batch.items()}
        with torch.no_grad():
            outputs = model(**batch)
        logits = outputs.logits
        probabilities = torch.nn.functional.softmax(logits, dim=-1)[:, 1]
        predictions = torch.argmax(logits, dim=-1)
        
        all_predictions.extend(predictions.cpu().numpy())
        all_labels.extend(batch["labels"].cpu().numpy())
        all_probabilities.extend(probabilities.detach().cpu().numpy())

    # Compute metrics
    accuracy = accuracy_score(all_labels, all_predictions)
    precision = precision_score(all_labels, all_predictions, average="weighted")
    recall = recall_score(all_labels, all_predictions, average="weighted")
    f1 = f1_score(all_labels, all_predictions, average="weighted")
    roc_auc = roc_auc_score(all_labels, all_probabilities)
    cm = confusion_matrix(all_labels, all_predictions)

    # Print metrics for the current epoch
    print(f"Epoch {epoch + 1} - Accuracy: {accuracy:.4f}, Precision: {precision:.4f}, Recall: {recall:.4f}, F1: {f1:.4f}, ROC AUC: {roc_auc:.4f}, Confusion Matrix: \n{cm}")
    
    # Set model back to training mode
    model.train()

# Evaluation on test set
model.eval()
all_labels = []
all_predictions = []
all_probabilities = []
for batch in test_loader:
    batch = {k: v.to(device) for k, v in batch.items()}
    with torch.no_grad():
        outputs = model(**batch)
    logits = outputs.logits
    probabilities = torch.nn.functional.softmax(logits, dim=-1)[:, 1]
    predictions = torch.argmax(logits, dim=-1)
    
    all_predictions.extend(predictions.cpu().numpy())
    all_labels.extend(batch["labels"].cpu().numpy())
    all_probabilities.extend(probabilities.detach().cpu().numpy())

# Compute final metrics
accuracy = accuracy_score(all_labels, all_predictions)
precision = precision_score(all_labels, all_predictions, average="weighted")
recall = recall_score(all_labels, all_predictions, average="weighted")
f1 = f1_score(all_labels, all_predictions, average="weighted")
roc_auc = roc_auc_score(all_labels, all_probabilities)
cm = confusion_matrix(all_labels, all_predictions)

# Stop the CodeCarbon tracker and calculate training time
tracker.stop()
training_time = (time.time() - start_time) / 60  # in minutes

print(f"Accuracy: {accuracy:.4f}, Precision: {precision:.4f}, Recall: {recall:.4f}, F1: {f1:.4f}, ROC AUC: {roc_auc:.4f}, Confusion Matrix: \n{cm}")

# Prepare data for saving
distilbert_data = {
    "Model": "DistilBERT with Pruning",
    "Dataset": "Amazon Polarity",
    "Evaluation Framework": "codecarbon",
    "Total Energy (kWh)": tracker.final_emissions_data.energy_consumed,
    "Total CO2 Emissions (kgCO2e)": tracker.final_emissions_data.emissions,
    "CPU Energy": tracker.final_emissions_data.cpu_energy,
    "GPU Energy": tracker.final_emissions_data.gpu_energy,
    "RAM Energy": tracker.final_emissions_data.ram_energy,
    "Emissions Rate": tracker.final_emissions_data.emissions_rate,
    "Training Time (minutes)": training_time,
    "Accuracy": accuracy,
    "Precision": precision,
    "Recall": recall,
    "F1": f1,
    "ROC AUC": roc_auc,
    "Confusion Matrix": "refer to terminal",
    "Number of Epochs": num_epochs
}

# Write data to Excel
wb = load_workbook("distilbertcompresults.xlsx")
ws = wb.active
ws.append(["", ""])
for key, value in distilbert_data.items():
    ws.append([key, value])

wb.save("distilbertcompresults.xlsx")
print("DistilBERT with Pruning data added to distilbertcompresults.xlsx")
